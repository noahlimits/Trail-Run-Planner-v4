#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_training_plan_v5.py
--------------------------------
Complete, self-contained script.

Changes vs prior:
- Evergreen plan = 12 weeks (84 days). Long runs forced to weekends when Firefighter = no.
- Weekly-hours scaling adjusts BOTH duration strings AND description text.
- Terrain type is a dropdown in Variables and alters sessions (road/flat removes downhill, mountainous adds more hills, etc.).
- Race block: distance (km) and elevation gain (m) are REQUIRED if a race date is given and they INFORM the plan:
    * Distance scales long-run caps, may add back-to-backs (for ≥70 km), and fueling-practice notes.
    * Elevation gain increases frequency/intensity of uphill/downhill work.
- Treadmill availability toggle (non-firefighter only) swaps Roche treadmill sessions appropriately.
- File saved to Desktop with timestamp: training_plan_YYYYMMDD_HHMM.xlsx
"""

import datetime as _dt
import re
import os
from pathlib import Path
from dataclasses import dataclass
from typing import List, Tuple, Optional

import pandas as _pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------- #
# ------------------------  Constants / Enums  ------------------------- #
# --------------------------------------------------------------------- #

TERRAIN_OPTIONS = ["Road/Flat", "Flat Trail", "Hilly Trail", "Mountainous/Skyrace"]

CATEGORY_HR = {
    "easy": ("<VT1",),
    "long": ("<VT1",),
    "roche": ("<VT1",),
    "threshold": (0.85, 0.90),
    "vo2": (0.90, 0.95),
    "speed": (0.95, 1.00),
    "strength": ("<VT1",),
    "recovery": ("<VT1",),
    "rest": ("rest",),
    "downhill": (0.90, 0.95),
}

CATEGORY_RPE = {
    "easy": "2–4",
    "long": "3–5",
    "roche": "3–5",
    "threshold": "5–7",
    "vo2": "8–9",
    "speed": "8–9",
    "strength": "3–5",
    "recovery": "1–2",
    "rest": "0",
    "downhill": "5–7",
}

DISTANCE_SUGGEST = {
    "10 km": "3–5",
    "21 km": "5–7",
    "42 km": "6–10",
    "50 km": "8–12",
    "70 km": "9–13",
    "100 km": "10–15",
}

# Distance-based scaling & logic
DISTANCE_LONG_CAP = {  # max long-run (unscaled) in minutes inside race block
    "short": 120,   # ≤30 km
    "mid":   180,   # 31–60 km
    "long":  240,   # 61–100 km
    "ultra": 300,   # >100 km (you can adjust)
}
# Back-to-back threshold
B2B_THRESHOLD_KM = 70

# Elevation gain thresholds (meters) to bump hill/downhill frequency
EG_LOW = 500
EG_MED = 1500
EG_HIGH = 2500

# --------------------------------------------------------------------- #
# ---------------------------  Data Model  ----------------------------- #
# --------------------------------------------------------------------- #

@dataclass
class ScheduleEntry:
    session: str
    description: str
    duration: str
    category: str


# --------------------------------------------------------------------- #
# ----------------------------  Utilities  ----------------------------- #
# --------------------------------------------------------------------- #

def _parse_date(text: str) -> _dt.date:
    return _dt.datetime.strptime(text, "%Y-%m-%d").date()


def _get_hr_range(category: str, hrmax: int, vt1: int) -> str:
    tpl = CATEGORY_HR.get(category, ("<VT1",))
    if tpl == ("rest",):
        return "-"
    if tpl[0] == "<VT1":
        return f"<{vt1} bpm"
    lo, hi = tpl
    lo_bpm = int(round(lo * hrmax))
    hi_bpm = int(round(hi * hrmax))
    return f"{lo_bpm}–{hi_bpm} bpm"


def _get_rpe(category: str) -> str:
    return CATEGORY_RPE.get(category, "3–5")


def _scale_duration(duration_str: str, category: str, scale: float) -> str:
    if category == "rest":
        return "-"
    if scale == 1.0:
        return duration_str

    nums = list(map(int, re.findall(r"\d+", duration_str)))
    if not nums:
        return duration_str

    def clamp(val, cat):
        if cat == "long":
            return max(60, min(val, 300))
        elif cat in ("threshold", "vo2", "speed", "downhill"):
            return max(20, min(val, 120))
        elif cat in ("easy", "roche", "strength", "recovery"):
            return max(20, min(val, 90))
        return val

    if len(nums) == 1:
        v = clamp(int(round(nums[0] * scale)), category)
        return f"{v} min"
    else:
        lo = clamp(int(round(nums[0] * scale)), category)
        hi = clamp(int(round(nums[1] * scale)), category)
        if lo == hi:
            return f"{lo} min"
        return f"{lo}-{hi} min"


def _update_description(description: str, scaled_duration: str) -> str:
    if "approx." in description:
        return description
    return f"{description} (approx. {scaled_duration})"


def _week_number(day_index: int) -> int:
    return day_index // 7 + 1


def _day_name(date: _dt.date) -> str:
    return date.strftime("%A")


def _parse_weekly_hours(text: str) -> float:
    if "-" in text:
        a, b = text.split("-", 1)
        try:
            return (float(a) + float(b)) / 2.0
        except ValueError:
            return 10.0
    try:
        return float(text)
    except ValueError:
        return 10.0


def _block_focus(week: int, include_base: bool) -> str:
    base_offset = 0 if include_base else 4
    if week <= 4 - base_offset:
        return "Base & Economy"
    elif week <= 8 - base_offset:
        return "Threshold & VO₂max"
    else:
        return "Speed-Endurance & Technical"


# --------------------------------------------------------------------- #
# --------------------- Terrain & Race Adjustments --------------------- #
# --------------------------------------------------------------------- #

def _adjust_for_terrain(session: str, description: str, category: str,
                        terrain: str, is_long: bool) -> Tuple[str, str, str]:
    s_low = session.lower()
    desc = description
    cat = category

    if terrain == "Hilly Trail":
        return session, description, category  # baseline

    if terrain == "Road/Flat":
        if "downhill" in s_low:
            session = "Neuromuscular Strides Session"
            desc = "60–75 min Z1 + 10×20 s fast strides on flat (no downhill reps)"
            cat = "speed"
        if ("uphill" in s_low or "hill" in s_low) and cat == "threshold":
            session = "Threshold Tempo (Flat)"
            desc = desc.replace("uphill", "flat").replace("hill", "flat")
            cat = "threshold"
        if is_long and "downhill" in desc.lower():
            desc = re.sub(r"\+.*downhill.*", "", desc, flags=re.IGNORECASE)
    elif terrain == "Flat Trail":
        if "downhill" in s_low:
            session = "Easy Run + Strides"
            desc = "60–75 min Z1 + 6×20 s strides (no downhill reps)"
            cat = "easy"
        if is_long and "downhill" in desc.lower():
            desc = re.sub(r"\+.*downhill.*", "", desc, flags=re.IGNORECASE)
    elif terrain == "Mountainous/Skyrace":
        if cat == "threshold" and "uphill" not in s_low:
            session = session.replace("Threshold", "Threshold Uphill")
            desc = desc if "uphill" in desc.lower() else desc + " (perform on sustained uphill)"
        if cat == "vo2" and "uphill" not in s_low:
            session = session.replace("VO₂max", "VO₂max Uphill")
            desc = desc + " (perform uphill if possible)"
        if is_long and "downhill" not in desc.lower():
            desc += " + 6×60 s downhill reps"

    return session, desc, cat


def _distance_bucket(km: int) -> str:
    if km <= 30:
        return "short"
    if km <= 60:
        return "mid"
    if km <= 100:
        return "long"
    return "ultra"


def _race_long_cap(km: int) -> int:
    return DISTANCE_LONG_CAP[_distance_bucket(km)]


def _apply_distance_elevation_logic(
    session: str,
    description: str,
    category: str,
    dist_km: int,
    elev_gain: int,
    day_index: int,
    long_cap: int,
    add_back_to_back: bool,
    fueling_required: bool,
) -> Tuple[str, str, str]:

    # Fueling practice insertion for long runs
    if category == "long" and fueling_required:
        if "fuel" not in description.lower():
            description += " (practice race fueling/hydration)"

    # Distance cap for long runs is enforced later via scaling; here we only change text if needed
    # Elevation gain: increase uphill/downhill specificity
    if elev_gain >= EG_HIGH:
        if category == "long" and "downhill" not in description.lower():
            description += " + 8×60 s downhill reps"
        if category == "threshold" and "uphill" not in description.lower():
            description += " (uphill focus)"

    elif elev_gain >= EG_MED:
        if category == "long" and "downhill" not in description.lower():
            description += " + 6×45 s downhill reps"

    # Back-to-back: if flagged and this is Saturday long, insert Sun medium-long in race builder (handled at sheet level by duplicating day? -> simpler: change Sunday easy to med-long)
    # We'll flag back-to-back outside (in generator) by altering Sunday easy run.
    return session, description, category


# --------------------------------------------------------------------- #
# -----------------------  Base / Race Schedules ----------------------- #
# --------------------------------------------------------------------- #

def _base_data() -> List[ScheduleEntry]:
    # 12 weeks × 7 days = 84 entries
    W = []

    # Weeks 1–4 Base & Economy
    W += [
        # Week 1
        ("Easy Run + Strides", "45–60 min Z1, finish 6×20 s strides", "45–60 min", "easy"),
        ("Threshold Hill Repeats 3×8 min", "WU 15', 3×8 min Z2 uphill, 2' jog, CD 10'", "60–75 min", "threshold"),
        ("Easy Run / Mobility", "50–60 min Z1 + mobility", "50–60 min", "easy"),
        ("Easy Run + Strength/Plyo", "40–50 min Z1 + heavy lifts ≥80%1RM; 3×8 plyo jumps", "75–90 min", "strength"),
        ("Easy Run", "45–60 min Z1", "45–60 min", "easy"),
        ("Long Run (Tech Trail)", "120–150 min Z1; add 6×30 s downhill strides", "120–150 min", "long"),
        ("Rest / Recovery", "Rest or 30 min easy cross", "-", "rest"),
        # Week 2
        ("Easy Run + Hill Sprints", "50–60 min Z1 + 6×10–15 s hill sprints", "50–60 min", "easy"),
        ("Threshold Tempo 20 min", "WU 15', 20' continuous Z2, CD 10'", "55–65 min", "threshold"),
        ("Easy Run / Mobility", "50–60 min Z1 + mobility", "50–60 min", "easy"),
        ("Easy Run + Strength/Plyo", "45–60 min Z1 + heavy lifts + plyo", "75–90 min", "strength"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("Long Run (Hilly)", "130–160 min Z1 on hilly trail", "130–160 min", "long"),
        ("Rest / Recovery", "Rest or gentle spin", "-", "rest"),
        # Week 3
        ("Easy Run + Strides", "50–60 min Z1, 6×20 s strides", "50–60 min", "easy"),
        ("Threshold Hill Repeats 5×5 min", "WU 15', 5×5 min uphill Z2, 2' jog, CD 10'", "65–80 min", "threshold"),
        ("Easy Run", "45–60 min Z1", "45–60 min", "easy"),
        ("Easy Run + Strength/Plyo", "40–50 min Z1 + lifts/plyo", "75–90 min", "strength"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("Long Run + Downhills", "140–170 min Z1 w/ 8×45 s downhill repeats", "140–170 min", "long"),
        ("Rest / Recovery", "Rest or yoga", "-", "rest"),
        # Week 4 (deload)
        ("Easy Run + Strides", "45–55 min Z1, 4×20 s strides", "45–55 min", "easy"),
        ("Light Threshold 2×10 min", "WU 10', 2×10' Z2 w/ 2' jog, CD 10'", "50–60 min", "threshold"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("Easy Run + Light Strength", "40–45 min Z1 + 1× heavy lifts", "60–70 min", "strength"),
        ("Easy Run", "40–50 min Z1", "40–50 min", "easy"),
        ("Long Run (Deload)", "100–120 min Z1", "100–120 min", "long"),
        ("Rest / Recovery", "Full rest", "-", "rest"),
    ]

    # Weeks 5–8 Threshold & VO2max
    W += [
        # Week 5
        ("Easy Run + Strides", "50–60 min Z1, 6×20 s strides", "50–60 min", "easy"),
        ("Threshold Hill Tempo 25 min", "WU 15', 25' uphill Z2 continuous, CD 10'", "60–75 min", "threshold"),
        ("Easy Run", "55–65 min Z1", "55–65 min", "easy"),
        ("VO₂max 4×3 min", "WU 15', 4×3' @90–95% HRmax, 3' jog, CD 10'", "60–70 min", "vo2"),
        ("Easy Run + Strength/Plyo", "40–50 min Z1 + lifts & plyo", "75–90 min", "strength"),
        ("Long Run (Tech Trail)", "150–180 min Z1, surges optional", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 6
        ("Easy Run + Hill Strides", "50–60 min Z1 + 6×20 s hill strides", "50–60 min", "easy"),
        ("Threshold 4×10 min", "WU 15', 4×10' Z2 w/ 3' jog, CD 10'", "75–90 min", "threshold"),
        ("Easy Run", "55–60 min Z1", "55–60 min", "easy"),
        ("VO₂max 3×4 min", "WU 15', 3×4' @90–95% HRmax, 3' jog, CD 10'", "55–65 min", "vo2"),
        ("Easy Run + Strength", "45–55 min Z1 + heavy lifts", "75–90 min", "strength"),
        ("Long Run + Downhills", "150–180 min Z1 w/ 6×60 s downhill reps", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 7
        ("Easy Run + Strides", "50–60 min Z1, 6×20 s strides", "50–60 min", "easy"),
        ("Threshold Mix 5×6 min", "WU 15', 5×6' Z2, 2' jog, CD 10'", "60–70 min", "threshold"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("VO₂max Uphill 4×3 min", "WU 15', 4×3' uphill @90–95%, 3' jog", "60–70 min", "vo2"),
        ("Easy Run + Strength/Plyo", "45–55 min Z1 + lifts/plyo", "75–90 min", "strength"),
        ("Long Run", "150–180 min Z1 technical", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 8 (deload)
        ("Easy Run + Strides", "45–55 min Z1, 4×20 s strides", "45–55 min", "easy"),
        ("Light Threshold 2×15 min", "WU 15', 2×15' Z2, CD 10'", "60–75 min", "threshold"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("VO₂max Short 2×3 min", "WU 15', 2×3' @90–95%, 3' jog, CD 10'", "45–55 min", "vo2"),
        ("Easy Run + Light Strength", "40–45 min Z1 + 1× lifts", "60–70 min", "strength"),
        ("Long Run (Deload)", "110–130 min Z1", "110–130 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]

    # Weeks 9–12 Speed-Endurance & Technical
    W += [
        # Week 9
        ("Easy Run + Strides", "50–60 min Z1, 6×20 s strides", "50–60 min", "easy"),
        ("Speed Endurance 6×1 min", "WU 15', 6×1' @~100% vVO₂, 2' jog, CD 10'", "50–60 min", "speed"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("Threshold Uphill 3×10 min", "WU 15', 3×10' Z2 uphill, 3' jog", "60–75 min", "threshold"),
        ("Easy Run + Strength/Plyo", "45–50 min Z1 + lifts/plyo", "75–90 min", "strength"),
        ("Long Run + Downhill Repeats", "140–170 min Z1 + 8×60 s downhill reps", "140–170 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 10
        ("Easy Run + Strides", "45–55 min Z1, 6×20 s strides", "45–55 min", "easy"),
        ("Speed Endurance 1×8×30-30", "WU 15', 8×(30\" fast/30\" float), CD 10'", "45–55 min", "speed"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("Threshold Tempo 20 min", "WU 15', 20' Z2 continuous, CD 10'", "55–65 min", "threshold"),
        ("Easy Run + Strength", "45–50 min Z1 + lifts", "75–90 min", "strength"),
        ("Long Run", "140–170 min Z1 technical", "140–170 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 11
        ("Easy Run + Strides", "45–55 min Z1, 6×20 s strides", "45–55 min", "easy"),
        ("Speed Endurance 8×45 s", "WU 15', 8×45\" fast / 90\" jog", "45–55 min", "speed"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("Threshold Combo 2×8 + 2×4 min", "WU 15', 2×8' Z2 + 2×4' upper Z2, CD", "60–70 min", "threshold"),
        ("Easy Run + Light Strength", "40–50 min Z1 + bodyweight plyo", "60–75 min", "strength"),
        ("Long Run (Shorten)", "120–140 min Z1", "120–140 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
        # Week 12
        ("Easy Run + Strides", "40–50 min Z1, 4×20 s strides", "40–50 min", "easy"),
        ("Speed Endurance 6×1 min", "WU 15', 6×1' fast / 2' jog, CD 10'", "45–55 min", "speed"),
        ("Easy Run", "40–50 min Z1", "40–50 min", "easy"),
        ("Threshold Uphill 2×10 min", "WU 15', 2×10' Z2 uphill, CD 10'", "50–60 min", "threshold"),
        ("Easy Run / Mobility", "40–50 min Z1 + mobility", "40–50 min", "easy"),
        ("Long Run (Recovery)", "100–120 min Z1", "100–120 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    return [ScheduleEntry(*row) for row in W]


def _race_data() -> List[ScheduleEntry]:
    # 10 weeks × 7 days = 70 entries (includes taper + race + short recovery)
    R = []
    # Week 1
    R += [
        ("Easy Run + Hill Strides", "60–75 min Z1 + 5×30 s hill strides", "60–75 min", "easy"),
        ("Threshold Uphill Tempo 25 min", "WU 15', 25' Z2 uphill, CD 10'", "60–75 min", "threshold"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("VO₂max 4×3 min", "WU 15', 4×3' @90–95% HRmax, 3' jog", "55–65 min", "vo2"),
        ("Easy Run + Strength", "45–55 min Z1 + lifts/plyo", "75–90 min", "strength"),
        ("Long Run + Downhill Repeats", "150–180 min Z1 + 6×60 s downhill reps", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 2
    R += [
        ("Easy Run + Strides", "55–65 min Z1 + 6×20 s strides", "55–65 min", "easy"),
        ("Threshold Mix 5×6 min", "WU 15', 5×6' Z2, 2' jog", "60–70 min", "threshold"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("VO₂max Uphill 3×4 min", "WU 15', 3×4' uphill @90–95%", "50–60 min", "vo2"),
        ("Easy Run + Strength", "45–55 min Z1 + lifts", "75–90 min", "strength"),
        ("Long Run (Tech)", "150–180 min Z1", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 3
    R += [
        ("Easy Run + Strides", "50–60 min Z1 + 6×20 s strides", "50–60 min", "easy"),
        ("Hill Beast 10/8/6/4/2 min", "WU 15', 10/8/6/4/2' uphill Z2, jog equal", "65–80 min", "threshold"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("VO₂max 4×3 min", "WU 15', 4×3' @90–95%, 3' jog", "55–65 min", "vo2"),
        ("Easy Run + Strength", "45–55 min Z1 + lifts", "75–90 min", "strength"),
        ("Long Run + Downhills", "160–190 min Z1 + 8×60 s downhill reps", "160–190 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 4
    R += [
        ("Easy Run + Strides", "50–60 min Z1 + 6×20 s strides", "50–60 min", "easy"),
        ("Threshold Uphill 3×10 min", "WU 15', 3×10' Z2 uphill", "60–75 min", "threshold"),
        ("Easy Run", "50–60 min Z1", "50–60 min", "easy"),
        ("Speed Endurance 1×8×30-30", "WU 15', 8×(30\" fast/30\" float), CD 10'", "45–55 min", "speed"),
        ("Easy Run + Light Strength", "40–50 min Z1 + bodyweight plyo", "60–75 min", "strength"),
        ("Long Run (Last Big)", "150–180 min Z1", "150–180 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 5
    R += [
        ("Easy Run + Strides", "45–55 min Z1 + 4×20 s strides", "45–55 min", "easy"),
        ("Threshold 2×15 min", "WU 15', 2×15' Z2, 3' jog", "60–75 min", "threshold"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("VO₂max 3×3 min", "WU 15', 3×3' @90–95%, 3' jog", "45–55 min", "vo2"),
        ("Easy Run + Strength", "40–50 min Z1 + light lifts", "60–75 min", "strength"),
        ("Long Run (Shorter)", "130–150 min Z1", "130–150 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 6
    R += [
        ("Easy Run + Strides", "45–55 min Z1 + 4×20 s strides", "45–55 min", "easy"),
        ("Speed Endurance 6×1 min", "WU 15', 6×1' fast / 2' jog", "45–55 min", "speed"),
        ("Easy Run", "45–55 min Z1", "45–55 min", "easy"),
        ("Threshold 20 min", "WU 15', 20' Z2, CD 10'", "55–65 min", "threshold"),
        ("Easy Run", "40–50 min Z1", "40–50 min", "easy"),
        ("Long Run (2 hr)", "120 min Z1", "120 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 7
    R += [
        ("Easy Run + Strides", "40–50 min Z1 + 4×20 s strides", "40–50 min", "easy"),
        ("VO₂max 2×3 min", "WU 15', 2×3' @90–95%, 3' jog", "40–50 min", "vo2"),
        ("Easy Run", "40–50 min Z1", "40–50 min", "easy"),
        ("Threshold 2×8 min", "WU 15', 2×8' Z2, CD 10'", "45–55 min", "threshold"),
        ("Easy Run", "40–45 min Z1", "40–45 min", "easy"),
        ("Long Run (90–100 min)", "90–100 min Z1", "90–100 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 8
    R += [
        ("Easy Run + Strides", "35–45 min Z1 + 4×20 s strides", "35–45 min", "easy"),
        ("Speed Endurance 6×30 s", "WU 10', 6×30\" fast / 60\" jog", "35–45 min", "speed"),
        ("Easy Run", "35–45 min Z1", "35–45 min", "easy"),
        ("Threshold 10 min", "WU 10', 10' Z2, CD 10'", "35–45 min", "threshold"),
        ("Easy Run", "30–40 min Z1", "30–40 min", "easy"),
        ("Long Run (70–80 min)", "70–80 min Z1", "70–80 min", "long"),
        ("Rest / Recovery", "Rest", "-", "rest"),
    ]
    # Week 9 (race week)
    R += [
        ("Easy Run", "30–40 min Z1", "30–40 min", "easy"),
        ("Hill Strides Only", "30–35 min Z1 + 6×20 s hill strides", "30–35 min", "easy"),
        ("Easy Run", "25–35 min Z1", "25–35 min", "easy"),
        ("Short VO₂ Touches 2×1 min", "WU 10', 2×1' fast / 2' jog", "25–35 min", "vo2"),
        ("Easy Run", "25–30 min Z1", "25–30 min", "easy"),
        ("Shakeout", "20–30 min very easy", "20–30 min", "easy"),
        ("Rest", "Race tomorrow", "-", "rest"),
    ]
    # Week 10 (race + recovery)
    R += [
        ("RACE DAY", "Race", "-", "rest"),
        ("Recovery / Off", "Full rest or 20 min walk", "-", "rest"),
        ("Easy Flush", "20–30 min Z1", "20–30 min", "easy"),
        ("Easy Run", "30–40 min Z1", "30–40 min", "easy"),
        ("Easy Run", "30–40 min Z1", "30–40 min", "easy"),
        ("Easy Run", "30–40 min Z1", "30–40 min", "easy"),
        ("Rest / Mobility", "Rest or yoga", "-", "rest"),
    ]
    return [ScheduleEntry(*row) for row in R]


# --------------------------------------------------------------------- #
# --------------------------  Core Builder  ---------------------------- #
# --------------------------------------------------------------------- #

def generate_plan(
    start_date: _dt.date,
    hrmax: int,
    vt1: int,
    vo2max: float,
    weekly_hours: str,
    shift_offset: int,
    race_date: Optional[_dt.date] = None,
    race_distance_km: Optional[int] = None,
    elevation_gain_m: Optional[int] = None,
    terrain_type: Optional[str] = None,
    include_base_block: bool = True,
    firefighter_schedule: bool = True,
    treadmill_available: bool = True,
) -> Tuple[_pd.DataFrame, _pd.DataFrame]:

    avg_hours = _parse_weekly_hours(weekly_hours)
    duration_scale = max(0.4, min(1.6, avg_hours / 10.0))

    terrain = terrain_type or "Hilly Trail"

    # ---------- Evergreen ----------
    base_sched = _base_data()
    if not include_base_block:
        base_sched = base_sched[28:]

    # Align start to Monday if not firefighter
    start_date_used = start_date if firefighter_schedule else start_date - _dt.timedelta(days=start_date.weekday())

    comp_rows = []
    for idx, entry in enumerate(base_sched):
        date = start_date_used + _dt.timedelta(days=idx)

        is_shift = False
        if firefighter_schedule:
            cyc = (idx + shift_offset) % 8
            is_shift = cyc in (0, 3)

        session = entry.session
        description = entry.description
        duration = entry.duration
        category = entry.category

        # Shift replacement
        if is_shift:
            if category in ("long", "threshold", "vo2", "speed", "downhill"):
                if category == "strength":
                    pass
                else:
                    session = "Roche Treadmill / Easy Run" if treadmill_available else "Easy Run + Hill Strides"
                    description = ("40–60 min Z2 uphill, incline to reach VT1"
                                   if treadmill_available else
                                   "60–75 min Z1 rolling + 6×30 s uphill strides")
                    duration = "40–75 min"
                    category = "roche" if treadmill_available else "easy"

        # Terrain adjust (only if not replaced)
        is_long = (category == "long")
        if not is_shift:
            session, description, category = _adjust_for_terrain(session, description, category, terrain, is_long)

        scaled_duration = _scale_duration(duration, category, duration_scale)
        if category != "rest":
            description = _update_description(description, scaled_duration)

        hr_range = _get_hr_range(category, hrmax, vt1)
        rpe = _get_rpe(category)

        comp_rows.append({
            "Week": _week_number(idx),
            "Date": date,
            "Day": _day_name(date),
            "Shift?": "Shift" if is_shift else "Off",
            "Session": session,
            "Description": description,
            "Duration": scaled_duration,
            "HR Target": hr_range,
            "RPE": rpe,
            "Block Focus": _block_focus(_week_number(idx), include_base_block)
        })

    comp_df = _pd.DataFrame(comp_rows)

    # ---------- Race ----------
    race_rows = []
    if race_date:
        if race_distance_km is None or elevation_gain_m is None:
            raise ValueError("Race distance and elevation gain must be provided when Race Date is set.")

        race_sched = _race_data()
        days_to_race = (race_date - start_date).days
        n_sessions = min(days_to_race + 1, len(race_sched))
        start_idx = len(race_sched) - n_sessions

        raw_start = race_date - _dt.timedelta(days=n_sessions - 1)
        if firefighter_schedule:
            aligned_start = raw_start
            delta_shift = (aligned_start - start_date).days
        else:
            aligned_start = raw_start - _dt.timedelta(days=raw_start.weekday())
            delta_shift = 0

        long_cap = _race_long_cap(race_distance_km)
        add_b2b = race_distance_km >= B2B_THRESHOLD_KM
        fueling_required = race_distance_km >= 42  # Marathon+ => practice fueling

        sunday_medium_long_indices = []  # track Sundays to convert for B2B

        for i in range(n_sessions):
            entry = race_sched[start_idx + i]
            date = aligned_start + _dt.timedelta(days=i) if not firefighter_schedule else raw_start + _dt.timedelta(days=i)

            if firefighter_schedule:
                cyc = (delta_shift + i + shift_offset) % 8
                is_shift = cyc in (0, 3)
            else:
                is_shift = False

            session = entry.session
            description = entry.description
            duration = entry.duration
            category = entry.category

            # Shift replacement
            if is_shift:
                if category in ("long", "threshold", "vo2", "speed", "downhill"):
                    if category == "strength":
                        pass
                    else:
                        session = "Roche Treadmill / Easy Run" if treadmill_available else "Easy Run + Hill Strides"
                        description = ("40–60 min Z2 uphill on treadmill"
                                       if treadmill_available
                                       else "60–75 min Z1 + 6×30 s hill strides")
                        duration = "40–75 min"
                        category = "roche" if treadmill_available else "easy"

            # Terrain adjust
            is_long = (category == "long")
            if not is_shift:
                session, description, category = _adjust_for_terrain(session, description, category, terrain, is_long)

            # Distance & elevation adjustments
            session, description, category = _apply_distance_elevation_logic(
                session, description, category, race_distance_km, elevation_gain_m,
                i, long_cap, add_b2b, fueling_required
            )

            scaled_duration = _scale_duration(duration, category, duration_scale)
            # Additional cap for long runs in race plan relative to distance
            if category == "long":
                nums = list(map(int, re.findall(r"\d+", scaled_duration)))
                if nums:
                    hi = max(nums)
                    if hi > long_cap:
                        # clamp to long_cap
                        scaled_duration = f"{min(nums)}-{long_cap} min" if len(nums) > 1 else f"{long_cap} min"

            if category != "rest":
                description = _update_description(description, scaled_duration)

            hr_range = _get_hr_range(category, hrmax, vt1)
            rpe = _get_rpe(category)

            # Track Sundays for possible B2B alteration
            if add_b2b and category == "long" and _day_name(date) == "Saturday":
                sunday_medium_long_indices.append(len(race_rows) + 1)  # next row index

            race_rows.append({
                "Date": date,
                "Day": _day_name(date),
                "Shift?": "Shift" if is_shift else "Off",
                "Session": session,
                "Description": description,
                "Duration": scaled_duration,
                "HR Target": hr_range,
                "RPE": rpe,
                "Block Focus": "Race Build"
            })

        # Convert following Sunday's easy run to Medium-Long if B2B
        if add_b2b:
            for idx_next in sunday_medium_long_indices:
                if idx_next < len(race_rows):
                    row = race_rows[idx_next]
                    if row["Day"] == "Sunday" and row["Session"].lower().startswith("easy"):
                        row["Session"] = "Medium-Long Run (B2B)"
                        row["Description"] = "90–120 min Z1 (back-to-back) (approx. 90–120 min)"
                        row["Duration"] = "90–120 min"
                        row["HR Target"] = f"<{vt1} bpm"
                        row["RPE"] = "3–5"

    race_df = _pd.DataFrame(race_rows)
    return comp_df, race_df


# --------------------------------------------------------------------- #
# ---------------------------  Save to XLSX  --------------------------- #
# --------------------------------------------------------------------- #

def save_plan_to_excel(comp_df: _pd.DataFrame,
                       race_df: _pd.DataFrame,
                       vars_dict: dict,
                       filename: str) -> None:
    wb = Workbook()
    ws_vars = wb.active
    ws_vars.title = "Variables"

    ws_vars.append(["Variable", "Value"])
    for k, v in vars_dict.items():
        ws_vars.append([k, v])

    # Terrain dropdown
    terrain_row = None
    for i, row in enumerate(ws_vars.iter_rows(values_only=True), start=1):
        if row and row[0] == "Terrain Type":
            terrain_row = i
            break
    if terrain_row:
        dv = DataValidation(type="list",
                            formula1=f'"{",".join(TERRAIN_OPTIONS)}"',
                            allow_blank=True,
                            showDropDown=True)
        ws_vars.add_data_validation(dv)
        dv.add(ws_vars[f"B{terrain_row}"])

    # Comprehensive
    ws_c = wb.create_sheet("Comprehensive Plan")
    ws_c.append(list(comp_df.columns))
    for r in comp_df.itertuples(index=False):
        ws_c.append(list(r))

    # Race
    ws_r = wb.create_sheet("Race Plan")
    if race_df.empty:
        ws_r.append(["(Enter a race date, distance & elevation in Variables and rerun the script)"])
    else:
        ws_r.append(list(race_df.columns))
        for r in race_df.itertuples(index=False):
            ws_r.append(list(r))

    wb.save(filename)


# --------------------------------------------------------------------- #
# -----------------------------  CLI Part  ----------------------------- #
# --------------------------------------------------------------------- #

def prompt_user() -> dict:
    print("Enter variables. Hit Enter for defaults where offered.\n")
    start = input("Start date (YYYY-MM-DD): ").strip()
    hrmax = input("Max HR (HRmax): ").strip()
    vt1 = input("VT1 (Aerobic threshold HR): ").strip()
    vo2 = input("VO2max (optional): ").strip()

    print("\nSuggested weekly training hours:")
    for k, v in DISTANCE_SUGGEST.items():
        print(f"  {k}: {v}")
    hrs = input("Weekly running time (hours, e.g. '8-12' or '6'): ").strip()

    base_choice = input("Include base-building block (y/n)? Default=y: ").strip().lower()
    ff_choice = input("Use firefighter shift schedule (y/n)? Default=y: ").strip().lower()

    tm_choice = ""
    if ff_choice in {"n", "no"}:
        tm_choice = input("Is a treadmill available (y/n)? Default=y: ").strip().lower()

    print("\nTerrain type options:")
    for i, opt in enumerate(TERRAIN_OPTIONS, 1):
        print(f"  {i}. {opt}")
    t_idx = input("Select terrain (1-4) or leave blank for 'Hilly Trail': ").strip()
    terrain = TERRAIN_OPTIONS[int(t_idx) - 1] if t_idx.isdigit() and 1 <= int(t_idx) <= 4 else "Hilly Trail"

    race = input("Race date (YYYY-MM-DD, optional): ").strip()
    rdist = input("Race distance in km (required if race date set): ").strip()
    egain = input("Elevation gain in m (required if race date set): ").strip()

    shift_off = input("Shift cycle offset (integer, default 0): ").strip()

    return {
        "Start Date": start,
        "Max HR (HRmax)": hrmax,
        "VT1": vt1,
        "VO2max": vo2,
        "Weekly Hours": hrs,
        "Include Base Block": base_choice,
        "Firefighter Schedule": ff_choice,
        "Treadmill Available": tm_choice,
        "Terrain Type": terrain,
        "Race Date": race,
        "Race Distance (km)": rdist,
        "Elevation Gain (m)": egain,
        "Shift Offset": shift_off,
    }


def main() -> None:
    vars_dict = prompt_user()

    start_date = _parse_date(vars_dict["Start Date"])
    hrmax = int(vars_dict["Max HR (HRmax)"])
    vt1 = int(vars_dict["VT1"])
    vo2 = float(vars_dict["VO2max"]) if vars_dict["VO2max"] else 0.0
    weekly_hours = vars_dict["Weekly Hours"]
    include_base = vars_dict["Include Base Block"].lower() not in {"n", "no"}
    firefighter = vars_dict["Firefighter Schedule"].lower() not in {"n", "no"}
    treadmill_available = vars_dict["Treadmill Available"].lower() not in {"n", "no"} if vars_dict["Treadmill Available"] else True
    terrain_type = vars_dict["Terrain Type"]

    race_date = _parse_date(vars_dict["Race Date"]) if vars_dict["Race Date"] else None
    race_distance = int(vars_dict["Race Distance (km)"]) if vars_dict["Race Distance (km)"] else None
    elevation_gain = int(vars_dict["Elevation Gain (m)"]) if vars_dict["Elevation Gain (m)"] else None
    shift_offset = int(vars_dict["Shift Offset"]) if vars_dict["Shift Offset"] else 0

    comp_df, race_df = generate_plan(
        start_date=start_date,
        hrmax=hrmax,
        vt1=vt1,
        vo2max=vo2,
        weekly_hours=weekly_hours,
        shift_offset=shift_offset,
        race_date=race_date,
        race_distance_km=race_distance,
        elevation_gain_m=elevation_gain,
        terrain_type=terrain_type,
        include_base_block=include_base,
        firefighter_schedule=firefighter,
        treadmill_available=treadmill_available,
    )

    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    default_name = f"training_plan_{stamp}.xlsx"
    desktop = Path.home() / "Desktop"
    outfile = (desktop / default_name) if desktop.exists() else Path.cwd() / default_name

    save_plan_to_excel(comp_df, race_df, vars_dict, str(outfile))
    print(f"\nSaved plan to: {outfile}\n")


if __name__ == "__main__":
    main()
